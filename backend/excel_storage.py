"""
Excel-based storage system for Budget Calculator
Stores all data in readable Excel format with multiple sheets
"""
import os
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_FILE = "budget_data.xlsx"
BACKUP_DIR = "backups"

def ensure_backup_dir():
    """Create backup directory if it doesn't exist"""
    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)

def create_excel_file():
    """Create a new Excel file - SIMPLIFIED: Only Income and Expense"""
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # SIMPLIFIED: Only Income and Expense sheets
    sheets = {
        'Income': ['ID', 'Date', 'Time', 'Category', 'Subcategory', 'Amount', 'Description', 'Created At', 'Updated At'],
        'Expense': ['ID', 'Date', 'Time', 'Category', 'Subcategory', 'Amount', 'Description', 'Created At', 'Updated At'],
        'Summary': ['Metric', 'Value', 'Last Updated']
    }
    
    # Create monthly sheets for current year and next year
    from datetime import datetime
    current_year = datetime.now().year
    months = ['January', 'February', 'March', 'April', 'May', 'June', 
              'July', 'August', 'September', 'October', 'November', 'December']
    
    for year in [current_year, current_year + 1]:
        for month in months:
            sheet_name = f"{month} {year}"
            sheets[sheet_name] = ['ID', 'Date', 'Time', 'Type', 'Category', 'Subcategory', 'Amount', 'Description', 'Created At']
    
    # Header styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name, headers in sheets.items():
        ws = wb.create_sheet(sheet_name)
        
        # Set headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    wb.save(EXCEL_FILE)
    return wb

def load_excel_file():
    """Load existing Excel file or create new one"""
    if os.path.exists(EXCEL_FILE):
        return load_workbook(EXCEL_FILE)
    else:
        return create_excel_file()

def get_sheet_data(wb, sheet_name: str) -> List[Dict]:
    """Get all data from a sheet as list of dictionaries"""
    if sheet_name not in wb.sheetnames:
        return []
    
    ws = wb[sheet_name]
    data = []
    
    # Skip header row
    if ws.max_row <= 1:
        return []  # Only headers, no data
    
    headers = [cell.value for cell in ws[1] if cell.value]
    
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_data = {}
        has_data = False
        
        for idx, cell in enumerate(row):
            if idx < len(headers) and headers[idx]:
                value = cell.value
                row_data[headers[idx]] = value
                # Check if this cell has actual data (not None, not empty string, not just whitespace)
                if value is not None and str(value).strip():
                    has_data = True
        
        # Only add rows that have at least one non-empty value
        if has_data:
            data.append(row_data)
    
    return data

def save_sheet_data(wb, sheet_name: str, data: List[Dict], headers: List[str]):
    """Save data to a sheet"""
    print(f"📄 save_sheet_data: {sheet_name}, {len(data)} records")
    
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
    else:
        ws = wb[sheet_name]
        # Clear existing data (keep headers)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    
    # Write data
    if data and len(data) > 0:
        print(f"✍️  Writing {len(data)} rows to {sheet_name}...")
        for row_num, row_data in enumerate(data, 2):
            row_values = []
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                row_values.append(value)
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            if row_num <= 3:  # Debug first few rows
                print(f"   Row {row_num}: {row_values[:5]}...")
    else:
        print(f"⚠️  No data to write to {sheet_name}")
    
    # Auto-adjust column widths
    if ws.max_row > 1:
        for col_num in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=row_num, column=col_num).value or ''))
                for row_num in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 30)
    
    print(f"💾 Saving workbook...")
    wb.save(EXCEL_FILE)
    print(f"✅ Workbook saved! {sheet_name} now has {ws.max_row} rows")

def backup_excel_file():
    """Create a timestamped backup of the Excel file"""
    ensure_backup_dir()
    if os.path.exists(EXCEL_FILE):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = os.path.join(BACKUP_DIR, f"budget_data_backup_{timestamp}.xlsx")
        import shutil
        shutil.copy2(EXCEL_FILE, backup_file)
        return backup_file
    return None

def load_all_data():
    """Load all data from Excel - SIMPLIFIED: Only Income and Expense"""
    wb = load_excel_file()
    
    print(f"📂 Loading data from Excel file: {EXCEL_FILE}")
    print(f"📋 Available sheets: {wb.sheetnames}")
    
    # Load Income and Expense sheets
    income_data = get_sheet_data(wb, 'Income')
    expense_data = get_sheet_data(wb, 'Expense')
    
    print(f"💰 Income sheet: {len(income_data)} records")
    print(f"💸 Expense sheet: {len(expense_data)} records")
    
    # Convert to transactions format for app compatibility
    transactions = []
    
    # Convert Income to transactions
    for income in income_data:
        transactions.append({
            'ID': income.get('ID'),
            'Date': income.get('Date'),
            'Time': income.get('Time'),
            'Type': 'Income',
            'Category': income.get('Category'),
            'Subcategory': income.get('Subcategory'),
            'Amount': income.get('Amount'),
            'Description': income.get('Description'),
            'Created At': income.get('Created At')
        })
    
    # Convert Expense to transactions
    for expense in expense_data:
        transactions.append({
            'ID': expense.get('ID'),
            'Date': expense.get('Date'),
            'Time': expense.get('Time'),
            'Type': 'Expense',
            'Category': expense.get('Category'),
            'Subcategory': expense.get('Subcategory'),
            'Amount': expense.get('Amount'),
            'Description': expense.get('Description'),
            'Created At': expense.get('Created At')
        })
    
    # IMPORTANT: DO NOT load from monthly sheets if main sheets are empty
    # Monthly sheets are only for organization when saving, not as a data source when fetching
    # If main Income/Expense sheets are empty, Excel is truly empty
    # This prevents confusion where Excel looks empty but fetch returns old data from monthly sheets
    if not transactions:
        print("ℹ️  Main Income and Expense sheets are empty. Excel file is empty.")
        print("   Monthly sheets are NOT checked - they are only for organization, not data storage.")
        print("   To add data: Use the app to add transactions, then click 'Update Excel'.")
    
    # Convert to expenses format (for Expense Calculator)
    expenses = []
    for expense in expense_data:
        expenses.append({
            'ID': expense.get('ID'),
            'Date': expense.get('Date'),
            'Time': expense.get('Time'),
            'Category': expense.get('Category'),
            'Subcategory': expense.get('Subcategory'),
            'Amount': expense.get('Amount'),
            'Description': expense.get('Description'),
            'Created At': expense.get('Created At')
        })
    
    print(f"📊 Returning: {len(transactions)} transactions, {len(expenses)} expenses")
    
    return {
        'transactions': transactions,
        'expenses': expenses,
        'savings': [],  # Removed - not needed
        'budgets': [],  # Removed - not needed
        'summary': get_sheet_data(wb, 'Summary')
    }

def save_transaction(transaction: Dict):
    """Save a transaction to Excel - saves to main sheet and monthly sheet"""
    wb = load_excel_file()
    transactions = get_sheet_data(wb, 'All Transactions')
    if not transactions:
        transactions = get_sheet_data(wb, 'Transactions')  # Backward compatibility
    
    # Generate ID if not present
    if not transaction.get('ID'):
        if transactions:
            max_id = max(int(t.get('ID', 0) or 0) for t in transactions)
            transaction['ID'] = max_id + 1
        else:
            transaction['ID'] = 1
    
    # Set timestamps
    now = datetime.now()
    if not transaction.get('Date'):
        transaction['Date'] = now.strftime('%Y-%m-%d')
    if not transaction.get('Time'):
        transaction['Time'] = now.strftime('%H:%M:%S')
    if not transaction.get('Created At'):
        transaction['Created At'] = now.isoformat()
    
    # Check if updating existing
    existing_idx = None
    for idx, t in enumerate(transactions):
        if t.get('ID') == transaction['ID']:
            existing_idx = idx
            break
    
    if existing_idx is not None:
        transactions[existing_idx] = transaction
    else:
        transactions.append(transaction)
    
    headers = ['ID', 'Date', 'Time', 'Type', 'Category', 'Subcategory', 'Amount', 'Description', 'Created At']
    
    # Save to main sheet and monthly sheet
    save_to_monthly_sheet(wb, [transaction], 'Transactions', headers)
    
    # Update summary
    update_summary(wb)
    
    return transaction

def update_summary(wb):
    """Update summary sheet - SIMPLIFIED: Only Income and Expense"""
    income_data = get_sheet_data(wb, 'Income')
    expense_data = get_sheet_data(wb, 'Expense')
    
    # If sheets don't exist, try old format
    if not income_data:
        all_transactions = get_sheet_data(wb, 'All Transactions') or get_sheet_data(wb, 'Transactions')
        income_data = [t for t in all_transactions if (t.get('Type') or '').lower() == 'income']
    
    if not expense_data:
        all_transactions = get_sheet_data(wb, 'All Transactions') or get_sheet_data(wb, 'Transactions')
        expense_data = [t for t in all_transactions if (t.get('Type') or '').lower() == 'expense']
        # Also get from Expenses sheet
        expenses_sheet = get_sheet_data(wb, 'All Expenses') or get_sheet_data(wb, 'Expenses')
        if expenses_sheet:
            expense_data.extend(expenses_sheet)
    
    income_total = sum(float(i.get('Amount', 0) or 0) for i in income_data)
    expense_total = sum(float(e.get('Amount', 0) or 0) for e in expense_data)
    balance = income_total - expense_total
    
    # Get or create Summary sheet
    if 'Summary' not in wb.sheetnames:
        ws = wb.create_sheet('Summary')
        headers = ['Metric', 'Value', 'Last Updated']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
    else:
        ws = wb['Summary']
        # Clear existing data (keep headers)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
    
    summary_data = [
        {'Metric': 'Total Income', 'Value': f'€{income_total:.2f}', 'Last Updated': datetime.now().isoformat()},
        {'Metric': 'Total Expenses', 'Value': f'€{expense_total:.2f}', 'Last Updated': datetime.now().isoformat()},
        {'Metric': 'Current Balance', 'Value': f'€{balance:.2f}', 'Last Updated': datetime.now().isoformat()},
        {'Metric': 'Total Income Records', 'Value': len(income_data), 'Last Updated': datetime.now().isoformat()},
        {'Metric': 'Total Expense Records', 'Value': len(expense_data), 'Last Updated': datetime.now().isoformat()},
        {'Metric': 'Last Updated', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'Last Updated': datetime.now().isoformat()},
    ]
    
    headers = ['Metric', 'Value', 'Last Updated']
    save_sheet_data(wb, 'Summary', summary_data, headers)

def get_month_sheet_name(date_str):
    """Get month sheet name from date string"""
    try:
        from datetime import datetime
        date_obj = datetime.strptime(date_str.split('T')[0], '%Y-%m-%d')
        months = ['January', 'February', 'March', 'April', 'May', 'June', 
                  'July', 'August', 'September', 'October', 'November', 'December']
        return f"{months[date_obj.month - 1]} {date_obj.year}"
    except:
        # Default to current month if parsing fails
        from datetime import datetime
        now = datetime.now()
        months = ['January', 'February', 'March', 'April', 'May', 'June', 
                  'July', 'August', 'September', 'October', 'November', 'December']
        return f"{months[now.month - 1]} {now.year}"

def save_to_monthly_sheet(wb, data, sheet_base_name, headers):
    """Save data to both main sheet (Income/Expense) and monthly sheet"""
    # Save to main Income or Expense sheet
    if sheet_base_name == 'Transactions':
        # Split into Income and Expense
        income_data = [d for d in data if (d.get('Type') or '').lower() == 'income']
        expense_data = [d for d in data if (d.get('Type') or '').lower() != 'income']
        if income_data:
            save_sheet_data(wb, 'Income', income_data, headers)
        if expense_data:
            save_sheet_data(wb, 'Expense', expense_data, headers)
    elif sheet_base_name == 'Income':
        save_sheet_data(wb, 'Income', data, headers)
    elif sheet_base_name == 'Expense':
        save_sheet_data(wb, 'Expense', data, headers)
    else:
        # Fallback to old naming
        save_sheet_data(wb, f'All {sheet_base_name}', data, headers)
    
    # Also save to monthly sheets
    monthly_data = {}
    for item in data:
        date_str = item.get('Date', item.get('Created At', ''))
        if date_str:
            month_sheet = get_month_sheet_name(date_str)
            if month_sheet not in monthly_data:
                monthly_data[month_sheet] = []
            monthly_data[month_sheet].append(item)
    
    # Save to each monthly sheet
    for month_sheet, month_items in monthly_data.items():
        # Get or create monthly sheet
        if month_sheet not in wb.sheetnames:
            ws = wb.create_sheet(month_sheet)
            # Add headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
        else:
            ws = wb[month_sheet]
            # Clear existing data (keep headers)
            ws.delete_rows(2, ws.max_row)
        
        # Write data
        for row_num, row_data in enumerate(month_items, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=row_num, column=col_num).value or ''))
                for row_num in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col_num)].width = min(max_length + 2, 30)
    
    wb.save(EXCEL_FILE)

def save_all_data(data: Dict):
    """Save all data to Excel - SIMPLIFIED: Only Income and Expense"""
    print(f"📝 save_all_data called with keys: {list(data.keys())}")
    print(f"📊 Transactions count: {len(data.get('transactions', []))}")
    print(f"📊 Expenses count: {len(data.get('expenses', []))}")
    
    wb = load_excel_file()
    
    # Create backup before updating
    backup_excel_file()
    
    total_records = 0
    
    # Combine transactions and expenses into Income/Expense
    income_records = []
    expense_records = []
    
    # Process transactions
    if 'transactions' in data and data['transactions']:
        for t in data['transactions']:
            try:
                date_str = t.get('Date') or t.get('date') or ''
                if date_str:
                    if 'T' in date_str:
                        date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    else:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    date_formatted = date_obj.strftime('%Y-%m-%d')
                    time_formatted = date_obj.strftime('%H:%M:%S')
                else:
                    date_formatted = datetime.now().strftime('%Y-%m-%d')
                    time_formatted = datetime.now().strftime('%H:%M:%S')
                
                record = {
                    'ID': t.get('ID') or t.get('id'),
                    'Date': date_formatted,
                    'Time': time_formatted,
                    'Category': t.get('Category') or t.get('category') or '',
                    'Subcategory': t.get('Subcategory') or t.get('subcategory') or '',
                    'Amount': float(t.get('Amount') or t.get('amount') or 0),
                    'Description': t.get('Description') or t.get('description') or '',
                    'Created At': t.get('Created At') or t.get('createdAt') or datetime.now().isoformat(),
                    'Updated At': datetime.now().isoformat()
                }
                if (t.get('Type') or t.get('type') or '').lower() == 'income':
                    income_records.append(record)
                else:
                    expense_records.append(record)
            except Exception as e:
                print(f"Error processing transaction: {e}")
                continue
    
    # Process expenses (all go to Expense sheet)
    if 'expenses' in data and data['expenses']:
        for e in data['expenses']:
            try:
                date_str = e.get('Date') or e.get('date') or ''
                if date_str:
                    if 'T' in date_str:
                        date_obj = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    else:
                        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    date_formatted = date_obj.strftime('%Y-%m-%d')
                    time_formatted = date_obj.strftime('%H:%M:%S')
                else:
                    date_formatted = datetime.now().strftime('%Y-%m-%d')
                    time_formatted = datetime.now().strftime('%H:%M:%S')
                
                record = {
                    'ID': e.get('ID') or e.get('id'),
                    'Date': date_formatted,
                    'Time': time_formatted,
                    'Category': e.get('Category') or e.get('category') or '',
                    'Subcategory': e.get('Subcategory') or e.get('subcategory') or '',
                    'Amount': float(e.get('Amount') or e.get('amount') or 0),
                    'Description': e.get('Description') or e.get('description') or '',
                    'Created At': e.get('Created At') or e.get('createdAt') or datetime.now().isoformat(),
                    'Updated At': datetime.now().isoformat()
                }
                expense_records.append(record)
            except Exception as err:
                print(f"Error processing expense: {err}")
                continue
    
    # Save Income sheet (even if empty, to clear old data)
    headers = ['ID', 'Date', 'Time', 'Category', 'Subcategory', 'Amount', 'Description', 'Created At', 'Updated At']
    print(f"💰 Income records to save: {len(income_records)}")
    print(f"💸 Expense records to save: {len(expense_records)}")
    
    # Always save Income sheet (use save_sheet_data to ensure proper saving)
    if income_records:
        print(f"💾 Saving {len(income_records)} income records to Income sheet...")
        save_sheet_data(wb, 'Income', income_records, headers)
        total_records += len(income_records)
        print(f"✅ Income records saved")
    else:
        print("⚠️  No income records to save - clearing Income sheet")
        # Clear Income sheet if no records
        if 'Income' in wb.sheetnames:
            ws = wb['Income']
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            wb.save(EXCEL_FILE)
    
    # Always save Expense sheet (use save_sheet_data to ensure proper saving)
    if expense_records:
        print(f"💾 Saving {len(expense_records)} expense records to Expense sheet...")
        save_sheet_data(wb, 'Expense', expense_records, headers)
        total_records += len(expense_records)
        print(f"✅ Expense records saved")
    else:
        print("⚠️  No expense records to save - clearing Expense sheet")
        # Clear Expense sheet if no records
        if 'Expense' in wb.sheetnames:
            ws = wb['Expense']
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)
            wb.save(EXCEL_FILE)
    
    # Update summary
    update_summary(wb)
    
    # IMPORTANT: Save the workbook after all updates
    try:
        wb.save(EXCEL_FILE)
        print(f"✅ Excel file saved successfully with {total_records} records")
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")
        raise
    
    return {"total_records": total_records, "status": "success", "income_count": len(income_records), "expense_count": len(expense_records)}

def clear_all_data():
    """Clear all data from Excel sheets (keeps structure and headers)"""
    wb = load_excel_file()
    
    print(f"🗑️  Clearing all data from Excel file: {EXCEL_FILE}")
    print(f"📋 Available sheets: {wb.sheetnames}")
    
    cleared_sheets = []
    
    # Clear Income sheet
    if 'Income' in wb.sheetnames:
        ws = wb['Income']
        rows_before = ws.max_row
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
            cleared_sheets.append('Income')
            print(f"✅ Cleared Income sheet: {rows_before - 1} rows deleted (now only headers)")
        else:
            print(f"ℹ️  Income sheet already empty")
    
    # Clear Expense sheet
    if 'Expense' in wb.sheetnames:
        ws = wb['Expense']
        rows_before = ws.max_row
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
            cleared_sheets.append('Expense')
            print(f"✅ Cleared Expense sheet: {rows_before - 1} rows deleted (now only headers)")
        else:
            print(f"ℹ️  Expense sheet already empty")
    
    # Clear all monthly sheets
    from datetime import datetime
    current_year = datetime.now().year
    months = ['January', 'February', 'March', 'April', 'May', 'June', 
              'July', 'August', 'September', 'October', 'November', 'December']
    
    monthly_cleared = 0
    monthly_rows_cleared = 0
    for year in [current_year - 1, current_year, current_year + 1]:
        for month in months:
            sheet_name = f"{month} {year}"
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_before = ws.max_row
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row)
                    monthly_cleared += 1
                    monthly_rows_cleared += (rows_before - 1)
    
    if monthly_cleared > 0:
        print(f"✅ Cleared {monthly_cleared} monthly sheets: {monthly_rows_cleared} total rows deleted")
    else:
        print(f"ℹ️  No monthly sheets to clear or already empty")
    
    # Clear Summary sheet data (keep structure)
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)
        # Add zero values
        ws.cell(row=2, column=1, value='Total Income')
        ws.cell(row=2, column=2, value=0)
        ws.cell(row=2, column=3, value=datetime.now().isoformat())
        ws.cell(row=3, column=1, value='Total Expense')
        ws.cell(row=3, column=2, value=0)
        ws.cell(row=3, column=3, value=datetime.now().isoformat())
        ws.cell(row=4, column=1, value='Balance')
        ws.cell(row=4, column=2, value=0)
        ws.cell(row=4, column=3, value=datetime.now().isoformat())
        cleared_sheets.append('Summary')
        print(f"✅ Cleared Summary sheet and reset to zeros")
    
    # IMPORTANT: Save the file
    try:
        wb.save(EXCEL_FILE)
        print(f"💾 Excel file saved successfully!")
        print(f"📊 Summary: Cleared {len(cleared_sheets)} main sheets ({', '.join(cleared_sheets)}) and {monthly_cleared} monthly sheets")
        print(f"✅ Excel file is now EMPTY (only headers remain)")
    except Exception as e:
        print(f"❌ Error saving Excel file: {e}")
        raise
    
    return {
        "status": "success", 
        "message": f"All data cleared from Excel! Cleared {len(cleared_sheets)} main sheets and {monthly_cleared} monthly sheets. Excel file is now empty."
    }



